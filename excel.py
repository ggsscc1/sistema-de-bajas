from tkinter import *
from tkinter import filedialog
from pathlib import Path
import tkinter as tk
from tkinter import ttk
import pandas as pd
import mysql.connector
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#Función para crear la ventana de consulta
def ventanaConsulta():
    # Crear una ventana tkinter
    global ventanaC
    ventanaC=tk.Tk()
    ventanaC.title("Resultados de la búsqueda")
    ventanaC.geometry("800x300")

    # Add a Scrollbar(horizontal)
    h=Scrollbar(ventanaC, orient='horizontal')
    h.pack(side=BOTTOM, fill='x')
    
    # Conexión a la base de datos
    conn = mysql.connector.connect(
    host='localhost',
    user='root',
    password='root',
    database='datosalumnosbajas'
    )

    cursor = conn.cursor()

    # Consulta SQL para obtener los datos
    query = "SELECT * FROM formulario"

    # Ejecutar la consulta SQL
    cursor.execute(query)

    # Obtener los resultados de la consulta
    results = cursor.fetchall()
   
    # Crear una tabla para mostrar los resultados
    global tabla_resultados
    tabla_resultados = ttk.Treeview(ventanaC, xscrollcommand=h.set)
    tabla_resultados['columns'] = ('ID', 'Clave', 'Nombre', 'Apellido paterno', 'Apellido materno',
                                   'Correo', 'Fecha de solicitud', 'Carrera', 'Generación', 'Tipo baja',
                                   'Motivo baja', 'Prepa origen', 'Materia difícil I', 'Materia difícil II',
                                   'Materia difícil III', 'Forma titulación', 'Fecha egel', 'Detalles baja', 'Empresa')

    tabla_resultados.column('#0', width=0, stretch=NO)
    tabla_resultados.column('ID', anchor=CENTER, width=40)
    tabla_resultados.column('Clave', anchor=CENTER, width=60)
    tabla_resultados.column('Nombre', anchor=CENTER, width=80)
    tabla_resultados.column('Apellido paterno', anchor=CENTER, width=110)
    tabla_resultados.column('Apellido materno', anchor=CENTER, width=110)
    tabla_resultados.column('Correo', anchor=CENTER, width=160)
    tabla_resultados.column('Fecha de solicitud', anchor=CENTER, width=120)
    tabla_resultados.column('Carrera', anchor=CENTER, width=130)
    tabla_resultados.column('Generación', anchor=CENTER, width=90)
    tabla_resultados.column('Tipo baja', anchor=CENTER, width=90)
    tabla_resultados.column('Motivo baja', anchor=CENTER, width=110)
    tabla_resultados.column('Prepa origen', anchor=CENTER, width=120)
    tabla_resultados.column('Materia difícil I', anchor=CENTER, width=130)
    tabla_resultados.column('Materia difícil II', anchor=CENTER, width=130)
    tabla_resultados.column('Materia difícil III', anchor=CENTER, width=130)
    tabla_resultados.column('Forma titulación', anchor=CENTER, width=130)
    tabla_resultados.column('Fecha egel', anchor=CENTER, width=120)
    tabla_resultados.column('Detalles baja', anchor=CENTER, width=110)
    tabla_resultados.column('Empresa', anchor=CENTER, width=100)

    tabla_resultados.heading('#0', text='', anchor=CENTER)
    tabla_resultados.heading('ID', text='ID', anchor=CENTER)
    tabla_resultados.heading('Clave', text='Clave', anchor=CENTER)
    tabla_resultados.heading('Nombre', text='Nombre', anchor=CENTER)
    tabla_resultados.heading('Apellido paterno', text='Apellido paterno', anchor=CENTER)
    tabla_resultados.heading('Apellido materno', text='Apellido materno', anchor=CENTER)
    tabla_resultados.heading('Correo', text='Correo', anchor=CENTER)
    tabla_resultados.heading('Fecha de solicitud', text='Fecha de solicitud', anchor=CENTER)
    tabla_resultados.heading('Carrera', text='Carrera', anchor=CENTER)
    tabla_resultados.heading('Generación', text='Generación', anchor=CENTER)
    tabla_resultados.heading('Tipo baja', text='Tipo baja', anchor=CENTER)
    tabla_resultados.heading('Motivo baja', text='Motivo baja', anchor=CENTER)
    tabla_resultados.heading('Prepa origen', text='Prepa origen', anchor=CENTER)
    tabla_resultados.heading('Materia difícil I', text='Materia difícil I', anchor=CENTER)
    tabla_resultados.heading('Materia difícil II', text='Materia difícil II', anchor=CENTER)
    tabla_resultados.heading('Materia difícil III', text='Materia difícil III', anchor=CENTER)
    tabla_resultados.heading('Forma titulación', text='Forma titulación', anchor=CENTER)
    tabla_resultados.heading('Fecha egel', text='Fecha egel', anchor=CENTER)
    tabla_resultados.heading('Detalles baja', text='Detalles baja', anchor=CENTER)
    tabla_resultados.heading('Empresa', text='Empresa', anchor=CENTER)

    # Insertar los resultados en la tabla
    for row in results:
        tabla_resultados.insert('', 'end', values=row)

    tabla_resultados.pack(expand=YES, fill=BOTH)

    # Attach the scrollbar with the text widget
    h.config(command=tabla_resultados.xview)

     # Agregar botón de exportar a Excel
    btn_exportar = Button(ventanaC, text="Exportar a Excel", command=exportar_a_excel)
    btn_exportar.pack()

    # Cerrar la conexión a la base de datos
    cursor.close()
    conn.close()
    ventanaC.mainloop()

# Función para exportar la tabla a un archivo de Excel
def exportar_a_excel():
    # Obtener los datos de la tabla
    datos = []
    for item in tabla_resultados.get_children():
        datos.append(tabla_resultados.item(item)['values'])

    # Crear un DataFrame de pandas con los datos
    df = pd.DataFrame(datos, columns=['ID', 'Clave', 'Nombre', 'Apellido paterno', 'Apellido materno',
                                      'Correo', 'Fecha de solicitud', 'Carrera', 'Generación', 'Tipo baja',
                                      'Motivo baja', 'Prepa origen', 'Materia difícil I', 'Materia difícil II',
                                      'Materia difícil III', 'Forma titulación', 'Fecha egel', 'Detalles baja', 'Empresa'])

    # Obtener el directorio y nombre del archivo para guardar
    directorio = filedialog.askdirectory()
    nombre_archivo = filedialog.asksaveasfilename(defaultextension=".xlsx")

    if directorio and nombre_archivo:
        # Crear un objeto Path para el directorio y el archivo
        directorio_path = Path(directorio)
        archivo_path = Path(nombre_archivo)

        # Combinar el directorio y el nombre de archivo correctamente
        archivo_guardar = directorio_path / archivo_path

        # Crear un objeto Workbook de openpyxl
        libro_excel = Workbook()
        hoja = libro_excel.active
        
        # Escribir los nombres de las columnas en la primera fila
        nombres_columnas = df.columns.tolist()
        hoja.append(nombres_columnas)

        # Convertir el DataFrame a una lista de listas
        datos_lista = df.values.tolist()

        # Escribir los datos en la hoja de Excel
        for row in datos_lista:
            hoja.append(row)

        # Cambiar el tamaño de las columnas
        for columna in hoja.columns:
            letra_columna = get_column_letter(columna[0].column)

            hoja.column_dimensions[letra_columna].width = 15


        hoja.column_dimensions['A'].width = 3
        # Guardar el archivo de Excel
        libro_excel.save(archivo_guardar)
        messagebox.showinfo("Exportar a Excel", "Los datos se exportaron correctamente.")
    else:
        messagebox.showwarning("Exportar a Excel", "Debes seleccionar un directorio y proporcionar un nombre de archivo.")


    """
    if directorio and nombre_archivo:
        # Crear un objeto Path para el directorio y el archivo
        directorio_path = Path(directorio)
        archivo_path = Path(nombre_archivo)

        # Combinar el directorio y el nombre de archivo correctamente
        archivo_guardar = directorio_path / archivo_path

        # Guardar el DataFrame en un archivo Excel
        
        df.to_excel(archivo_guardar, index=False)
        messagebox.showinfo("Exportar a Excel", "Los datos se exportaron correctamente.")
    else:
        messagebox.showwarning("Exportar a Excel", "Debes seleccionar un directorio y proporcionar un nombre de archivo.")   
    """
